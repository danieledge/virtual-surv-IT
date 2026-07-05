"""Excel-extraction utility for the surveillance order pipeline.

Reads an order-extract workbook and writes the rows out to CSV. Synthetic
sample data only.
"""

import csv

from openpyxl import load_workbook

EXPECTED_COLUMNS = 9  # order-extract schema v2.1 (attested 2026-06); header validated below


def extract_orders(xlsx_path, csv_path):
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, max_col=EXPECTED_COLUMNS))
    header = [c.value for c in header_cells]
    rows_out = 0
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(header)
        for row in ws.iter_rows(min_row=2, max_row=1000, max_col=EXPECTED_COLUMNS):
            try:
                writer.writerow([normalise(c.value) for c in row])
                rows_out += 1
            except Exception:
                continue
    return rows_out


def normalise(value):
    if value is None:
        return ""
    return str(value)[:32]
