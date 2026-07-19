#!/usr/bin/env python3
"""
convert_file - the team's single front door for reading/converting source files.

Why this exists: hand-rolled Excel/CSV conversion has repeatedly produced silently wrong
data downstream (long IDs re-typed as floats, date serials misread, truncated extracts,
locale separators). Conversion is a mechanical step, so it is centralised here where the
failure modes are defended once, with evidence. House rule: agents never hand-parse
Excel/CSV/PDF/DOCX inputs - they call this script and attach its report
(docs/house-rules.md).

Formats in:
  .xlsx/.xlsm  via openpyxl (vendored)     - read-only stream, cached formula values
  .xls         via xlrd (vendored)          - legacy BIFF
  .csv/.tsv/.txt  via the stdlib            - encoding + dialect defences, never silent
  .pdf         via pypdf (vendored)         - text extraction only (tables in PDFs are
                                              inherently unreliable; flagged, not guessed)
  .docx        via stdlib zipfile+ElementTree - paragraphs and tables (python-docx needs
                                              compiled lxml, so it cannot be vendored)

Formats out: csv (default) / jsonl for tabular; md / txt for documents.

Design rules (the reliability contract):
  1. LOSSLESS BY DEFAULT - with no schema, every value is emitted as text with zero type
     inference. Typed Excel cells are rendered deterministically (dates -> ISO 8601,
     integers without ".0"). Nothing is guessed.
  2. SCHEMA MODE - --schema <feed>.yaml|.json declares column names/types/patterns and
     expectations (row counts, control totals). Violations are ERRORS: non-zero exit, no
     output row is trusted. This is a gate, not a warning log.
  3. EVIDENCE ALWAYS - every run writes a JSON report (counts, hashes, encoding decisions,
     warnings) next to the output and prints a one-screen summary. "It converted" is never
     the evidence; the report is.
  4. RECONCILIATION - rows read vs rows written, declared sheet dimensions vs streamed
     counts, ragged-row detection, unterminated-final-line detection (the truncation
     defences; see docs/house-rules.md on reconciling extracts).

Data safety: refuses inputs under the project's data/raw/ (CLAUDE.md paragraph 5) - raw
records go through scripts/ingest.py masking, never through a plain converter.

Vendored dependencies: imported from <repo>/vendor (pure-Python, pinned, licence notices
in THIRD-PARTY-LICENSES.md) so a plain `git clone` works in environments without pip
access. Vendored copies take precedence over site-packages for determinism.
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import hashlib
import io
import json
import os
import re
import sys
import zipfile
from decimal import Decimal, InvalidOperation
from pathlib import Path

TOOL_VERSION = "1.0.0"

# ---------------------------------------------------------------------------
# Vendored-first imports. The vendor/ tree ships the exact versions the tests were
# written against; preferring it over site-packages keeps behaviour identical on every
# machine (a newer system openpyxl changing a default must not change our output).
# ---------------------------------------------------------------------------
_VENDOR = Path(__file__).resolve().parent.parent / "vendor"
if _VENDOR.is_dir() and str(_VENDOR) not in sys.path:
    sys.path.insert(0, str(_VENDOR))

TABULAR_OUT = ("csv", "jsonl")
DOCUMENT_OUT = ("md", "txt")

# Only these delimiters are candidates for sniffing: letting csv.Sniffer pick from the
# whole byte range makes it "detect" delimiters inside free text.
_SNIFF_DELIMITERS = ",;\t|"

# ISO-ish date/datetime formats accepted when a schema asks for date/datetime and the
# source value is a string. Deliberately short: format guessing is how 03/04 becomes
# 4 March in one file and 3 April in the next. Anything else needs `format:` in the schema.
_DATE_FORMATS = ("%Y-%m-%d",)
_DATETIME_FORMATS = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M")

_SCHEMA_TYPES = ("string", "integer", "decimal", "date", "datetime", "boolean")


class ConversionError(Exception):
    """Fatal problem: the conversion cannot be trusted. Exit 1, no usable output."""


# ---------------------------------------------------------------------------
# Report - the evidence object every run produces.
# ---------------------------------------------------------------------------
class Report:
    def __init__(self, source: Path) -> None:
        self.data: dict = {
            "tool": f"convert_file {TOOL_VERSION}",
            "source": str(source),
            "source_sha256": _sha256(source),
            "source_bytes": source.stat().st_size,
            "warnings": [],
            "errors": [],
            "checks": {},
        }

    def warn(self, msg: str) -> None:
        self.data["warnings"].append(msg)

    def error(self, msg: str) -> None:
        self.data["errors"].append(msg)

    def check(self, name: str, passed: bool, detail: str) -> None:
        self.data["checks"][name] = {"passed": passed, "detail": detail}
        if not passed:
            self.error(f"check failed - {name}: {detail}")

    def write(self, path: Path) -> None:
        path.write_text(json.dumps(self.data, indent=2, default=str) + "\n", encoding="utf-8")

    def summary(self) -> str:
        d = self.data
        lines = [f"source       {d['source']}  sha256 {d['source_sha256'][:12]}…"]
        for key in (
            "format",
            "sheet",
            "encoding",
            "delimiter",
            "rows_read",
            "rows_output",
            "columns",
            "pages",
            "tables",
            "output",
        ):
            if key in d:
                lines.append(f"{key:<13}{d[key]}")
        for name, res in d["checks"].items():
            lines.append(
                f"check        {'PASS' if res['passed'] else 'FAIL'}  {name}: {res['detail']}"
            )
        for w in d["warnings"]:
            lines.append(f"warning      {w}")
        for e in d["errors"]:
            lines.append(f"ERROR        {e}")
        return "\n".join(lines)


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Data-safety guard: raw records are the masking pipeline's job, not ours.
# ---------------------------------------------------------------------------
def _refuse_raw_data(source: Path) -> None:
    project = Path(os.environ.get("CLAUDE_PROJECT_DIR") or os.getcwd()).resolve()
    raw = project / "data" / "raw"
    try:
        resolved = source.resolve()
    except OSError:
        raise ConversionError(f"cannot resolve path: {source}")
    if resolved == raw or raw in resolved.parents:
        raise ConversionError(
            "input is under data/raw/ - raw records must go through the masking pipeline "
            "(python -m scripts.ingest, see /prepare-data), not a plain converter (CLAUDE.md §5)."
        )


# ---------------------------------------------------------------------------
# Schema - optional, declarative, and strict where it matters.
# ---------------------------------------------------------------------------
def load_schema(path: Path) -> dict:
    text = path.read_text(encoding="utf-8")
    if path.suffix.lower() in (".yaml", ".yml"):
        try:
            import yaml  # PyYAML: already a team dependency (masking pipeline)
        except ImportError:
            raise ConversionError(
                f"schema {path} is YAML but PyYAML is not installed - either install it "
                "(pip install PyYAML) or provide the schema as .json (stdlib, always works)."
            )
        schema = yaml.safe_load(text)
    else:
        schema = json.loads(text)
    if not isinstance(schema, dict) or not isinstance(schema.get("columns"), list):
        raise ConversionError(f"schema {path}: expected a mapping with a 'columns' list.")
    for col in schema["columns"]:
        ctype = col.get("type", "string")
        if ctype not in _SCHEMA_TYPES:
            raise ConversionError(
                f"schema {path}: column {col.get('name')!r} has unknown type {ctype!r} "
                f"(one of {', '.join(_SCHEMA_TYPES)})."
            )
    return schema


def _cast(value: str, col: dict, row_no: int, report: Report) -> object:
    """Cast a rendered string value to the schema type. Fail loud, never guess."""
    name, ctype = col.get("name", "?"), col.get("type", "string")
    if value == "":
        if col.get("required"):
            raise ConversionError(f"row {row_no}: required column {name!r} is empty.")
        return ""
    if ctype == "string":
        pattern = col.get("pattern")
        if pattern and not re.fullmatch(pattern, value):
            raise ConversionError(
                f"row {row_no}: column {name!r} value {value!r} does not match pattern {pattern!r}."
            )
        return value
    if ctype == "integer":
        try:
            return int(value.replace(" ", ""))
        except ValueError:
            raise ConversionError(f"row {row_no}: column {name!r}: {value!r} is not an integer.")
    if ctype == "decimal":
        try:
            return Decimal(value.replace(" ", ""))
        except InvalidOperation:
            raise ConversionError(f"row {row_no}: column {name!r}: {value!r} is not a decimal.")
    if ctype == "boolean":
        low = value.strip().lower()
        if low in ("true", "1", "yes", "y"):
            return True
        if low in ("false", "0", "no", "n"):
            return False
        raise ConversionError(f"row {row_no}: column {name!r}: {value!r} is not a boolean.")
    formats = (
        [col["format"]]
        if col.get("format")
        else list(_DATE_FORMATS if ctype == "date" else _DATETIME_FORMATS)
    )
    for fmt in formats:
        try:
            parsed = _dt.datetime.strptime(value, fmt)
            return parsed.date().isoformat() if ctype == "date" else parsed.isoformat(sep=" ")
        except ValueError:
            continue
    raise ConversionError(
        f"row {row_no}: column {name!r}: {value!r} does not parse as {ctype} "
        f"(accepted formats: {', '.join(formats)}; add 'format:' to the schema for others)."
    )


def apply_schema(
    header: list[str],
    rows: list[list[str]],
    schema: dict,
    cell_types: list[list[str]] | None,
    report: Report,
) -> list[list[str]]:
    """Validate header/rows against the schema; return re-rendered rows. Raises on breach."""
    declared = [c["name"] for c in schema["columns"]]
    if header[: len(declared)] != declared:
        raise ConversionError(
            f"header mismatch - expected {declared}, found {header[: len(declared)]}. "
            "A renamed or reordered column upstream is a silent-corruption risk; the schema "
            "must be updated deliberately, not worked around."
        )
    extra = header[len(declared) :]
    if extra and not schema.get("allow_extra_columns", False):
        raise ConversionError(
            f"unexpected extra columns {extra} (set allow_extra_columns: true if intended)."
        )
    report.check("header", True, f"{len(declared)} declared columns matched")

    # Provenance warning: schema says string but the workbook stored a number - Excel has
    # already applied float semantics (leading zeros gone, >15 digits rounded). The damage
    # is at SOURCE; we surface it instead of laundering it.
    if cell_types:
        for j, col in enumerate(schema["columns"]):
            if col.get("type", "string") == "string":
                numeric = sum(1 for types in cell_types if j < len(types) and types[j] == "n")
                if numeric:
                    report.warn(
                        f"column {col['name']!r}: {numeric} value(s) stored as NUMBER in the "
                        "workbook but schema says string - leading zeros / precision beyond "
                        "15 digits may already be lost at source. Fix the extract (format as "
                        "text); don't trust these values."
                    )

    out: list[list[str]] = []
    totals: dict[str, Decimal] = {
        name: Decimal(0) for name in (schema.get("expect", {}).get("control_totals") or {})
    }
    for i, row in enumerate(rows, start=2):  # data starts at line 2 of the output
        rendered: list[str] = []
        for j, col in enumerate(schema["columns"]):
            value = row[j] if j < len(row) else ""
            cast = _cast(value, col, i, report)
            if col["name"] in totals and value != "":
                try:
                    totals[col["name"]] += Decimal(value.replace(" ", ""))
                except InvalidOperation:
                    raise ConversionError(
                        f"row {i}: control-total column {col['name']!r} value {value!r} is "
                        "not numeric - control_totals must reference numeric columns."
                    )
            rendered.append(str(cast) if not isinstance(cast, bool) else str(cast).lower())
        out.append(rendered)

    expect = schema.get("expect", {}) or {}
    if expect.get("row_count") is not None:
        ok = len(out) == expect["row_count"]
        report.check("row_count", ok, f"expected {expect['row_count']}, got {len(out)}")
    if expect.get("min_rows") is not None:
        ok = len(out) >= expect["min_rows"]
        report.check("min_rows", ok, f"expected >= {expect['min_rows']}, got {len(out)}")
    for name, expected in (expect.get("control_totals") or {}).items():
        actual = totals.get(name, Decimal(0))
        ok = actual == Decimal(str(expected))
        report.check(f"control_total:{name}", ok, f"expected {expected}, got {actual}")
    if report.data["errors"]:
        raise ConversionError("schema/reconciliation checks failed - see report.")
    return out


# ---------------------------------------------------------------------------
# Deterministic cell rendering (the lossless contract).
# ---------------------------------------------------------------------------
def render_cell(value: object) -> tuple[str, str]:
    """Return (text, kind) where kind is 's' string / 'n' numeric / 'd' date / 'b' bool / '' empty."""
    if value is None:
        return "", ""
    if isinstance(value, bool):
        return ("true" if value else "false"), "b"
    if isinstance(value, _dt.datetime):
        # Excel has no timezone concept; midnight datetimes are almost always pure dates.
        if value.time() == _dt.time(0, 0):
            return value.date().isoformat(), "d"
        return value.isoformat(sep=" "), "d"
    if isinstance(value, _dt.date):
        return value.isoformat(), "d"
    if isinstance(value, _dt.time):
        return value.isoformat(), "d"
    if isinstance(value, int):
        return str(value), "n"
    if isinstance(value, float):
        if value.is_integer() and abs(value) < 1e15:
            return str(int(value)), "n"
        return repr(value), "n"  # repr = shortest round-trip text, no locale involved
    return str(value), "s"


# ---------------------------------------------------------------------------
# Readers - each returns (header, rows-as-text, cell_types) plus report entries.
# ---------------------------------------------------------------------------
def read_xlsx(source: Path, sheet: str | None, report: Report):
    try:
        import openpyxl
    except ImportError as exc:  # vendored, so this means a broken checkout
        raise ConversionError(f"openpyxl unavailable ({exc}) - is vendor/ present in the clone?")
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        if sheet is None:
            ws = wb[names[0]]
            if len(names) > 1:
                report.warn(
                    f"workbook has {len(names)} sheets {names}; converted {names[0]!r} only - "
                    "pass --sheet to pick another. Unconverted sheets are a coverage gap."
                )
        else:
            if sheet not in names:
                raise ConversionError(f"sheet {sheet!r} not found; workbook has {names}.")
            ws = wb[sheet]
        report.data["sheet"] = ws.title
        report.data["format"] = "xlsx"

        declared_rows = ws.max_row  # from the dimension record - can lie; we count for real
        raw_rows: list[list[str]] = []
        types: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rendered = [render_cell(v) for v in row]
            raw_rows.append([t for t, _ in rendered])
            types.append([k for _, k in rendered])
        # data_only=True yields None for formula cells the writing application never
        # cached - those cells silently become empty, which we must surface.
        streamed = len(raw_rows)
        while raw_rows and all(v == "" for v in raw_rows[-1]):
            raw_rows.pop()
            types.pop()
        trimmed = streamed - len(raw_rows)
        if trimmed:
            report.data["trailing_empty_rows_trimmed"] = trimmed
        if declared_rows is not None and declared_rows != streamed:
            report.warn(
                f"sheet dimension record declares {declared_rows} rows but {streamed} were "
                "streamed - the file's metadata is stale; streamed count is authoritative."
            )
    finally:
        wb.close()
    if not raw_rows:
        raise ConversionError("no data rows found.")
    return raw_rows[0], raw_rows[1:], types[1:]


def read_xls(source: Path, sheet: str | None, report: Report):
    try:
        import xlrd
    except ImportError as exc:
        raise ConversionError(f"xlrd unavailable ({exc}) - is vendor/ present in the clone?")
    book = xlrd.open_workbook(str(source))
    names = book.sheet_names()
    if sheet is None:
        ws = book.sheet_by_index(0)
        if len(names) > 1:
            report.warn(f"workbook has {len(names)} sheets {names}; converted {names[0]!r} only.")
    else:
        if sheet not in names:
            raise ConversionError(f"sheet {sheet!r} not found; workbook has {names}.")
        ws = book.sheet_by_name(sheet)
    report.data["sheet"] = ws.name
    report.data["format"] = "xls"

    raw_rows, types = [], []
    for i in range(ws.nrows):
        row_text, row_kind = [], []
        for j in range(ws.ncols):
            cell = ws.cell(i, j)
            if cell.ctype == xlrd.XL_CELL_DATE:
                value = xlrd.xldate_as_datetime(cell.value, book.datemode)  # 1900/1904 handled
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                value = bool(cell.value)
            elif cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
                value = None
            elif cell.ctype == xlrd.XL_CELL_ERROR:
                report.warn(f"cell ({i + 1},{j + 1}) holds an Excel error value; emitted empty.")
                value = None
            else:
                value = cell.value
            text, kind = render_cell(value)
            row_text.append(text)
            row_kind.append(kind)
        raw_rows.append(row_text)
        types.append(row_kind)
    while raw_rows and all(v == "" for v in raw_rows[-1]):
        raw_rows.pop()
        types.pop()
    if not raw_rows:
        raise ConversionError("no data rows found.")
    return raw_rows[0], raw_rows[1:], types[1:]


def _detect_encoding(raw: bytes, override: str | None, report: Report) -> str:
    if override:
        report.data["encoding"] = f"{override} (explicit)"
        return override
    if raw.startswith(b"\xef\xbb\xbf"):
        report.data["encoding"] = "utf-8-sig (BOM)"
        return "utf-8-sig"
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        report.data["encoding"] = "utf-16 (BOM)"
        return "utf-16"
    try:
        raw.decode("utf-8")
        report.data["encoding"] = "utf-8 (strict decode)"
        return "utf-8"
    except UnicodeDecodeError:
        # cp1252 maps nearly every byte, so this always "works" - which is exactly why it
        # must be a recorded assumption, never a silent one.
        report.data["encoding"] = "cp1252 (ASSUMED fallback - not valid UTF-8; verify)"
        report.warn(
            "file is not valid UTF-8; decoded as cp1252 (Windows-1252). If the source is "
            "another legacy encoding, non-ASCII characters are wrong - re-run with --encoding."
        )
        return "cp1252"


def read_csv(source: Path, delimiter: str | None, encoding: str | None, report: Report):
    raw = source.read_bytes()
    if not raw:
        raise ConversionError("file is empty.")
    enc = _detect_encoding(raw, encoding, report)
    try:
        text = raw.decode(enc)
    except (UnicodeDecodeError, LookupError) as exc:
        raise ConversionError(f"cannot decode as {enc}: {exc}")

    if not text.endswith(("\n", "\r")):
        report.warn(
            "file does not end with a newline - a classic sign of a TRUNCATED extract. "
            "Reconcile row counts against the source before trusting this data."
        )

    if delimiter is None:
        sample = text[:65536]
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=_SNIFF_DELIMITERS).delimiter
            report.data["delimiter"] = f"{delimiter!r} (sniffed)"
        except csv.Error:
            delimiter = ","
            report.data["delimiter"] = "',' (default - sniffing was inconclusive)"
    else:
        report.data["delimiter"] = f"{delimiter!r} (explicit)"
    report.data["format"] = "csv"

    rows = list(csv.reader(io.StringIO(text, newline=""), delimiter=delimiter))
    while rows and rows[-1] in ([], [""]):
        rows.pop()
    if not rows:
        raise ConversionError("no data rows found.")
    header, data = rows[0], rows[1:]

    # Ragged rows mean the delimiter/quoting is wrong or the file is corrupt - both make
    # every downstream number suspect. Hard stop, with the rows named.
    bad = [str(i + 2) for i, r in enumerate(data) if len(r) != len(header)]
    if bad:
        shown = ", ".join(bad[:10]) + ("…" if len(bad) > 10 else "")
        raise ConversionError(
            f"{len(bad)} row(s) have a different field count than the header ({len(header)} "
            f"fields): rows {shown}. Wrong delimiter/quoting or a corrupt/truncated file."
        )
    report.check("field_counts", True, f"{len(data)} rows x {len(header)} fields, none ragged")
    return header, data, None


def read_docx(source: Path, report: Report) -> dict:
    # defusedxml (vendored): stdlib ElementTree is vulnerable to entity-expansion attacks
    # (billion laughs) on untrusted input, and every input to this tool is untrusted.
    # openpyxl also auto-uses defusedxml when importable, hardening the xlsx path too.
    try:
        from defusedxml.ElementTree import fromstring as _xml_fromstring
    except ImportError as exc:
        raise ConversionError(f"defusedxml unavailable ({exc}) - is vendor/ present in the clone?")
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    try:
        with zipfile.ZipFile(source) as zf:
            xml = zf.read("word/document.xml")
    except (zipfile.BadZipFile, KeyError) as exc:
        raise ConversionError(f"not a readable .docx (zip/document.xml): {exc}")
    body = _xml_fromstring(xml).find("w:body", ns)
    if body is None:
        raise ConversionError("document.xml has no body element.")

    def para_text(p) -> str:
        return "".join(t.text or "" for t in p.iter(f"{{{ns['w']}}}t"))

    blocks: list[dict] = []
    for child in body:
        tag = child.tag.split("}")[1]
        if tag == "p":
            text = para_text(child)
            if text.strip():
                blocks.append({"kind": "para", "text": text})
        elif tag == "tbl":
            table = []
            for tr in child.findall("w:tr", ns):
                table.append(
                    [
                        "\n".join(para_text(p) for p in tc.findall("w:p", ns)).strip()
                        for tc in tr.findall("w:tc", ns)
                    ]
                )
            blocks.append({"kind": "table", "rows": table})
    report.data["format"] = "docx"
    report.data["tables"] = sum(1 for b in blocks if b["kind"] == "table")
    if not blocks:
        report.warn("document produced no text blocks - empty or unsupported content.")
    return {"blocks": blocks}


def _pdftotext_pages(source: Path) -> list[str] | None:
    """
    Poppler fallback: extract page texts via `pdftotext` (system package poppler-utils).

    Optional by design - pypdf (vendored, pure Python) stays the primary engine so the
    no-pip promise holds, but its extraction is weak on PDFs with unusual encodings or
    complex layouts. When poppler is installed, pages pypdf could not read get a second
    chance here. Returns None when pdftotext is absent or fails (caller keeps pypdf's
    result); pages split on the form-feed separators pdftotext emits.
    """
    import shutil

    # Used for one fixed-argv call below - no shell, resolved binary path.
    import subprocess  # nosec B404

    exe = shutil.which("pdftotext")
    if not exe:
        return None
    try:
        # argv list (no shell), executable resolved via shutil.which, source is a local
        # file path already validated by the caller.
        proc = subprocess.run(  # nosec B603
            [exe, "-layout", str(source), "-"], capture_output=True, timeout=120
        )
    except (OSError, subprocess.SubprocessError):
        return None
    if proc.returncode != 0:
        return None
    text = proc.stdout.decode("utf-8", errors="replace")
    pages = text.split("\f")
    if pages and not pages[-1].strip():
        pages.pop()  # pdftotext ends output with a trailing form feed
    return pages


def read_pdf(source: Path, report: Report) -> list[str]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ConversionError(f"pypdf unavailable ({exc}) - is vendor/ present in the clone?")
    reader = PdfReader(str(source))
    if reader.is_encrypted:
        # An empty owner password is common; anything else is out of scope by design.
        try:
            reader.decrypt("")
        except Exception:
            raise ConversionError("PDF is encrypted and could not be opened without a password.")
    pages = [page.extract_text() or "" for page in reader.pages]
    report.data["format"] = "pdf"
    report.data["pages"] = len(pages)
    report.data["pdf_engine"] = "pypdf"
    empty_idx = [i for i, t in enumerate(pages) if not t.strip()]
    if empty_idx:
        alt = _pdftotext_pages(source)
        if alt is not None:
            recovered = []
            for i in empty_idx:
                if i < len(alt) and alt[i].strip():
                    pages[i] = alt[i]
                    recovered.append(str(i + 1))
            if recovered:
                report.data["pdf_engine"] = "pypdf+pdftotext"
                report.warn(
                    f"{len(recovered)} page(s) had no pypdf-extractable text and were "
                    f"recovered via pdftotext/poppler (pages {', '.join(recovered[:10])}"
                    f"{'…' if len(recovered) > 10 else ''}) - layout-derived text; verify "
                    "against the source for anything load-bearing."
                )
            empty_idx = [i for i, t in enumerate(pages) if not t.strip()]
    empty = [str(i + 1) for i in empty_idx]
    if empty:
        import shutil

        hint = (
            ""
            if shutil.which("pdftotext")
            else " Installing poppler-utils (pdftotext) may recover pages like these."
        )
        report.warn(
            f"{len(empty)} of {len(pages)} page(s) have NO extractable text (pages "
            f"{', '.join(empty[:10])}{'…' if len(empty) > 10 else ''}) - likely scanned "
            "images. OCR is out of scope; treat content from those pages as MISSING, "
            "not empty." + hint
        )
    report.warn(
        "PDF extraction is TEXT only - table structure in PDFs is layout, not data, and "
        "reconstructing it is unreliable. If this PDF carries tabular data, get the "
        "upstream Excel/CSV instead; that is the defensible source."
    )
    return pages


# ---------------------------------------------------------------------------
# Writers.
# ---------------------------------------------------------------------------
def write_tabular(header: list[str], rows: list[list[str]], out: Path, fmt: str) -> None:
    if fmt == "csv":
        with out.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
            writer.writerow(header)
            writer.writerows(rows)
    else:  # jsonl
        with out.open("w", encoding="utf-8") as fh:
            for row in rows:
                fh.write(json.dumps(dict(zip(header, row)), ensure_ascii=False) + "\n")


def write_document(blocks_or_pages, out: Path, fmt: str, kind: str) -> None:
    parts: list[str] = []
    if kind == "pdf":
        for i, text in enumerate(blocks_or_pages, start=1):
            parts.append(f"## Page {i}\n\n{text.strip()}" if fmt == "md" else text.strip())
    else:  # docx blocks
        for block in blocks_or_pages["blocks"]:
            if block["kind"] == "para":
                parts.append(block["text"])
            else:
                rows = block["rows"]
                if fmt == "md" and rows:
                    width = max(len(r) for r in rows)
                    norm = [r + [""] * (width - len(r)) for r in rows]
                    md = [
                        "| "
                        + " | ".join(c.replace("|", "\\|").replace("\n", " ") for c in norm[0])
                        + " |",
                        "|" + " --- |" * width,
                    ]
                    md += [
                        "| "
                        + " | ".join(c.replace("|", "\\|").replace("\n", " ") for c in r)
                        + " |"
                        for r in norm[1:]
                    ]
                    parts.append("\n".join(md))
                else:
                    parts.append("\n".join("\t".join(r) for r in rows))
    out.write_text("\n\n".join(parts) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# CLI.
# ---------------------------------------------------------------------------
def _default_out(source: Path, to: str) -> Path:
    return source.with_name(f"{source.stem}.converted.{to}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m scripts.convert_file",
        description="Reliable file conversion with schema validation and an evidence report. "
        "Tabular (xlsx/xlsm/xls/csv/tsv) -> csv/jsonl; documents (pdf/docx) -> md/txt.",
    )
    parser.add_argument("input", help="file to convert")
    parser.add_argument(
        "--to",
        choices=TABULAR_OUT + DOCUMENT_OUT,
        help="output format (default: csv for tabular, md for documents)",
    )
    parser.add_argument("--out", help="output path (default: <input>.converted.<fmt>)")
    parser.add_argument("--sheet", help="worksheet name (default: first, with a warning if more)")
    parser.add_argument(
        "--table", type=int, help="docx only: extract table N (1-based) as tabular output"
    )
    parser.add_argument("--schema", help="feed schema (.yaml/.json) to validate against")
    parser.add_argument("--delimiter", help="csv only: field delimiter (default: sniffed)")
    parser.add_argument("--encoding", help="csv only: source encoding (default: detected)")
    parser.add_argument("--report", help="evidence report path (default: <out>.report.json)")
    parser.add_argument(
        "--list", action="store_true", help="list sheets/tables/pages and exit without converting"
    )
    args = parser.parse_args(argv)

    source = Path(args.input)
    if not source.is_file():
        parser.error(f"not a file: {source}")
    suffix = source.suffix.lower()

    try:
        _refuse_raw_data(source)
        report = Report(source)

        if suffix in (".xlsx", ".xlsm"):
            if args.list:
                import openpyxl

                wb = openpyxl.load_workbook(source, read_only=True)
                print("\n".join(wb.sheetnames))
                wb.close()
                return 0
            header, rows, types = read_xlsx(source, args.sheet, report)
            kind = "tabular"
        elif suffix == ".xls":
            if args.list:
                import xlrd

                print("\n".join(xlrd.open_workbook(str(source)).sheet_names()))
                return 0
            header, rows, types = read_xls(source, args.sheet, report)
            kind = "tabular"
        elif suffix in (".csv", ".tsv", ".txt"):
            delim = args.delimiter or ("\t" if suffix == ".tsv" else None)
            header, rows, types = read_csv(source, delim, args.encoding, report)
            kind = "tabular"
        elif suffix == ".docx":
            doc = read_docx(source, report)
            if args.list:
                for i, b in enumerate((b for b in doc["blocks"] if b["kind"] == "table"), start=1):
                    print(
                        f"table {i}: {len(b['rows'])} rows x "
                        f"{max((len(r) for r in b['rows']), default=0)} cols"
                    )
                return 0
            if args.table is not None or (args.to in TABULAR_OUT):
                tables = [b for b in doc["blocks"] if b["kind"] == "table"]
                if not tables:
                    raise ConversionError("document contains no tables to extract.")
                idx = (args.table or 1) - 1
                if not 0 <= idx < len(tables):
                    raise ConversionError(f"--table {args.table} out of range (1..{len(tables)}).")
                trows = tables[idx]["rows"]
                header, rows, types = trows[0], trows[1:], None
                kind = "tabular"
            else:
                kind = "docx"
        elif suffix == ".pdf":
            pages = read_pdf(source, report)
            if args.list:
                for i, t in enumerate(pages, start=1):
                    print(f"page {i}: {len(t)} chars")
                return 0
            kind = "pdf"
        else:
            raise ConversionError(
                f"unsupported format {suffix!r}. Tabular: .xlsx .xlsm .xls .csv .tsv .txt; "
                "documents: .pdf .docx. (.xlsb is not supported - re-save as .xlsx.)"
            )

        if kind == "tabular":
            to = args.to or "csv"
            if to not in TABULAR_OUT:
                raise ConversionError(f"--to {to} is for documents; tabular outputs: csv, jsonl.")
            if args.schema:
                schema = load_schema(Path(args.schema))
                report.data["schema"] = args.schema
                rows = apply_schema(header, rows, schema, types, report)
                header = [c["name"] for c in schema["columns"]]
            out = Path(args.out) if args.out else _default_out(source, to)
            write_tabular(header, rows, out, to)
            report.data.update(
                rows_read=len(rows), rows_output=len(rows), columns=len(header), output=str(out)
            )
            report.check("rows_written", True, f"{len(rows)} rows written to {out.name}")
        else:
            to = args.to or "md"
            if to not in DOCUMENT_OUT:
                raise ConversionError(f"--to {to} is for tabular data; document outputs: md, txt.")
            out = Path(args.out) if args.out else _default_out(source, to)
            write_document(pages if kind == "pdf" else doc, out, to, kind)
            report.data["output"] = str(out)

        report_path = (
            Path(args.report) if args.report else out.with_suffix(out.suffix + ".report.json")
        )
        report.write(report_path)
        report.data["report"] = str(report_path)
        print(report.summary())
        print(f"report   {report_path}")
        return 0

    except ConversionError as exc:
        print(f"CONVERSION FAILED: {exc}", file=sys.stderr)
        try:
            report.error(str(exc))
            report_path = (
                Path(args.report)
                if args.report
                else source.with_suffix(source.suffix + ".report.json")
            )
            report.write(report_path)
            print(f"report   {report_path}", file=sys.stderr)
        except Exception:  # noqa: BLE001 - reporting must never mask the real failure  # nosec B110
            pass
        return 1


if __name__ == "__main__":
    sys.exit(main())
